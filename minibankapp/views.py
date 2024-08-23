# -*- coding: utf-8 -*-

import json
import datetime
import decimal
from openpyxl import Workbook
from openpyxl.styles import (Border, Side, PatternFill, Font)
from django.http.response import HttpResponse as HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.mixins import (LoginRequiredMixin, PermissionRequiredMixin)
from django.contrib.auth.models import Permission
from django.views.generic import (View, CreateView, ListView, UpdateView, DeleteView, FormView)
from django.core.exceptions import ValidationError
from django.db.utils import DataError
from django.db import transaction
from django.urls import (reverse, reverse_lazy)
from django.db.models import ProtectedError
from .forms import (
                    CustomerForm,
                    NewAccountTypeForm, UpdateAccountTypeForm,
                    NewAccountForm, UpdateAccountForm,
                    UpdateParameterForm,
                    CreateOperationForm)

from .models import (CustomerModel, AccountModel, OperationModel, AccountTypeModel, ParameterModel)
from .functions import SelectCustomerListView


""" Custom Permission """
class CustomPermission(Permission):
    
    class Meta:
        default_permissions = ()
        permissions = (
            ("standard_role", "Standard role"),
            ("extended_role", "Extended role"))


""" Dashboard """
class MainDasboard(LoginRequiredMixin, View):

    def get(self, request):
        return render(request, 'minibankapp/dashboard.html')


""" Parameter """
class ParameterUpdateView(LoginRequiredMixin, UpdateView):
    form_class = UpdateParameterForm
    template_name = 'minibankapp/updateparameter.html'
    success_url = reverse_lazy('minibankapp:dashboard')

    def get_object(self):
        return ParameterModel.objects.get()
    

""" Customer """
class CustomerCreateView(LoginRequiredMixin, CreateView):
    form_class = CustomerForm
    template_name = 'minibankapp/newcustomer.html'
    
    def form_valid(self, form):
        instance = form.save(commit=False)
        instance.Created_employee = self.request.user
        try:
            instance.full_clean()
            instance.save()
            return redirect(reverse('minibankapp:newcustomer_done', args=[instance.pk]))
        except:
            return render(self.request, self.template_name, {
                                                            'form': form,
                                                            'error_message': 'Something went wrong ...'})
    
    def form_invalid(self, form):
        for field in form.errors:
            form[field].field.widget.attrs['class'] += ' errorfield'
        return super().form_invalid(form)


class CustomerCreateDoneView(LoginRequiredMixin, View):

    def get(self, request, *args, **kwargs):
        return render(request, 'minibankapp/newcustomer_done.html', {'pk': self.kwargs['customer']})


class CustomerListView(LoginRequiredMixin, SelectCustomerListView):
    template_name = 'minibankapp/viewcustomer.html'
    paginate_by = 10

    def get_queryset(self):
        if self.request.GET.get('criteria'):
            return self.get_queryset_customer('criteria')
        else:
            return CustomerModel.objects.filter().order_by('-pk')


class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    form_class = CustomerForm
    template_name = 'minibankapp/updatecustomer.html'
    success_url = reverse_lazy('minibankapp:listcustomer')

    def get_object(self):
        return CustomerModel.objects.get(pk=self.kwargs['customer'])
    
    def form_invalid(self, form):
        for field in form.errors:
            form[field].field.widget.attrs['class'] += ' errorfield'
        return super().form_invalid(form)


class CustomerDeleteView(LoginRequiredMixin, DeleteView):
    template_name = 'minibankapp/deletecustomer.html'
    
    def get_object(self):
        return CustomerModel.objects.get(pk=self.kwargs['customer'])
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            return redirect('/listcustomer/')
        except ProtectedError as error_description:
            return render(request,'minibankapp/error.html', {'error_message': error_description})


class SelectCustomerAccountListView(LoginRequiredMixin, SelectCustomerListView):
    template_name = 'minibankapp/selectcustomer_account.html'
    paginate_by = 10

    def get_queryset(self):
        if self.request.GET.get('criteria'):
            return self.get_queryset_customer('criteria')
        else:
            return CustomerModel.objects.filter().order_by('-pk')


""" Account """
class AccountListView(LoginRequiredMixin, ListView):
    template_name = 'minibankapp/viewaccount.html'

    def get_queryset(self):
        results = AccountModel.objects.filter(FK_Id_customer=self.kwargs['customer']).order_by('-pk')
        if results.exists():
            return results
        else:
            return AccountModel.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pk_customer'] = self.kwargs['customer']
        return context


class AccountCreateView(LoginRequiredMixin, CreateView):
    form_class = NewAccountForm
    template_name = 'minibankapp/newaccount.html'

    def form_valid(self, form):
        instance = form.save(commit=False)
        instance.Created_employee = self.request.user
        instance.FK_Id_customer_id = self.kwargs['customer']
        try:
            instance.Free_balance = instance.Debit
            instance.full_clean()
            instance.save()
            return redirect(reverse('minibankapp:listaccount', args=[self.kwargs['customer']]))
        except ValidationError as error_message:
            return render(self.request, self.template_name, {
                                                            'form': form,
                                                            'pk_customer': self.kwargs['customer'], 
                                                            'error_message': error_message.messages[0]})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        dt = {}
        instance = AccountTypeModel.objects.all().values_list('Id_account_type', 'Percent')
        for i in instance:
            dt[i[0]] = str(i[1])
        context['pk_customer'] = self.kwargs['customer']
        context['dataJSON'] = json.dumps(dt)
        return context

    
class AccountUpdateView(LoginRequiredMixin, UpdateView):
    form_class = UpdateAccountForm
    template_name = 'minibankapp/updateaccount.html'

    def get_object(self):
        return AccountModel.objects.get(pk=self.kwargs['account'])

    def get_success_url(self):
        return f"/listaccount/{self.kwargs['customer']}/"
          
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['Nr_IBAN'] = AccountModel.objects.get(Id_account=self.kwargs['account']).Number_IBAN
        return context
   
    def form_valid(self, form):
        # Updating free balance based on debit
        instance = form.save(commit=False)
        var_balance = AccountModel.objects.get(Id_account=self.kwargs['account']).Balance
        instance.Free_balance = instance.Debit + var_balance
        try:
            instance.full_clean()
            instance.save()
        except ValidationError as error_message:
            return render(self.request, self.template_name, {
                                                            'form': form,
                                                            'error_message': error_message.messages[0]})
        return super().form_valid(form)


""" Generating IBAN number """
class AccountGenerateUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'minibankapp.extended_role'

    def get_object(self):
        return AccountModel.objects.get(pk=self.kwargs['account'])
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        country_code = ParameterModel.objects.get().Country_code
        bank_number = ParameterModel.objects.get().Bank_number
        subaccount = AccountTypeModel.objects.get(Id_account_type=self.object.FK_Id_account_type.property_get_pk).Subaccount
        customer = str(self.kwargs['customer'])
        account = str(self.kwargs['account'])
        prefix_zero = ''
        while len(customer) + len(account) + len(prefix_zero) < 12:
            prefix_zero = prefix_zero + '0'
        # Creating IBAN and save
        self.object.Number_IBAN = country_code + bank_number + subaccount + account + prefix_zero + customer
        try:
            self.object.full_clean()
            self.object.save()
            return redirect(reverse('minibankapp:listaccount', args=[self.kwargs['customer']]))
        except (DataError, ProtectedError) as error_description:
            return render(request,'minibankapp/error.html', {'error_message': error_description})


""" Interest counting """
class AccountInterestUpdateView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    permission_required = 'minibankapp.extended_role'

    def get(self, request):
        return render(request, 'minibankapp/interest_confirm.html')
    
    def post(self, request):
        counter = 0
        data = AccountModel.objects.filter(Balance__gt = 0, Percent__gt = 0)
        if data.exists():
            try:
                with transaction.atomic():
                    for instance in data:
                        interest = round(instance.Balance * (instance.Percent / 100),2)
                        new_balance = instance.Balance + interest
                        new_free_balance = new_balance + instance.Debit
                        # Updating balance & free balance
                        AccountModel.objects.filter(Id_account=instance.Id_account).update(Balance=new_balance, Free_balance=new_free_balance)
                        # New object in OperationModel
                        OperationModel.objects.create(
                                                        Type_operation = 3,
                                                        Value_operation = interest,
                                                        Balance_after_operation = new_balance,
                                                        Operation_employee = self.request.user,
                                                        FK_Id_account = instance)
                        counter += 1
            except Exception as error_description:
                return render(request,'minibankapp/error.html', {'error_message': error_description})
        msg = 'Interest for ' + str(counter) + ' account(s) has been recounted.'
        return render(request, 'minibankapp/interest_done.html', {'msg': msg})


""" AccountType """
class AccountTypeCreateView(LoginRequiredMixin, CreateView):
    form_class = NewAccountTypeForm
    template_name = 'minibankapp/newaccounttype.html'
    success_url = reverse_lazy('minibankapp:listaccounttype')


class AccountTypeListView(LoginRequiredMixin, ListView):
    template_name = 'minibankapp/viewaccounttype.html'

    def get_queryset(self):
        if self.request.GET.get('criteria'):
            criterias = self.request.GET.get('criteria')
            results = AccountTypeModel.objects.filter(Id_account_type__istartswith=criterias).order_by('-pk')
            if results.exists():
                return results
            else:
                return AccountTypeModel.objects.none()
        return AccountTypeModel.objects.filter().order_by('-pk')


class AccountTypeUpdateView(LoginRequiredMixin, UpdateView):
    form_class = UpdateAccountTypeForm
    template_name = 'minibankapp/updateaccounttype.html'
    success_url = reverse_lazy('minibankapp:listaccounttype')

    def get_object(self):
        return AccountTypeModel.objects.get(pk=self.kwargs['accounttype'])


class AccountTypeDeleteView(LoginRequiredMixin, DeleteView):

    def get_object(self):
        return AccountTypeModel.objects.get(pk=self.kwargs['accounttype'])

    # Without confirmation
    def get(self, request, *args, **kwargs):
        return self.delete(request, *args, **kwargs)
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.delete()
            return redirect('/listaccounttype/')
        except ProtectedError as error_description:
            return render(request,'minibankapp/error.html', {'error_message': error_description})


""" Operation """
class OperationCreateView(LoginRequiredMixin, CreateView):
    form_class = CreateOperationForm
    template_name = 'minibankapp/newoperation.html'

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        self.var_free_balance = AccountModel.objects.get(pk=self.kwargs['account']).Free_balance
        self.var_balance = AccountModel.objects.get(pk=self.kwargs['account']).Balance
        self.var_debit = AccountModel.objects.get(pk=self.kwargs['account']).Debit
        self.var_iban = AccountModel.objects.get(pk=self.kwargs['account']).Number_IBAN
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['free_balance'] = self.var_free_balance
        context['balance'] = self.var_balance
        context['debit'] = self.var_debit
        context['nr_iban'] = self.var_iban
        context['pk_customer'] = self.kwargs['customer']
        return context
    
    def form_valid(self, form):     
        try:            
            with transaction.atomic():
                instance = form.save(commit=False)
                # Operation type
                if instance.Type_operation == 2:
                    instance.Value_operation = instance.Value_operation * (-1)
                # Set up balance after transaction
                var_balance_after_operation = AccountModel.objects.get(pk=self.kwargs['account']).Balance + instance.Value_operation
                instance.Balance_after_operation = var_balance_after_operation
                # Other data
                instance.Operation_employee = self.request.user
                instance.FK_Id_account_id = self.kwargs['account']
                # Updating balance & free balance for AccountModel
                var_free_balance_after_operation = var_balance_after_operation + self.var_debit
                record = AccountModel.objects.get(pk=self.kwargs['account'])
                record.Balance = var_balance_after_operation
                record.Free_balance = var_free_balance_after_operation
                record.full_clean()
                record.save()
                instance.full_clean()
                instance.save()
                return redirect(reverse('minibankapp:selectacount_operation', args=[self.kwargs['customer']]))
        except ValidationError as error_message:
            self.var_free_balance = AccountModel.objects.get(pk=self.kwargs['account']).Free_balance
            for content in error_message:
                item = content[1]
                error_message = item[0]
            return render(self.request, self.template_name, {
                                                                'form': form,
                                                                'nr_iban': self.var_iban,
                                                                'pk_customer': self.kwargs['customer'],
                                                                'balance': self.var_balance,
                                                                'debit': self.var_debit,
                                                                'free_balance': self.var_free_balance,
                                                                'error_message': error_message})

    def form_invalid(self, form):
        for field in form.errors:
            form[field].field.widget.attrs['class'] += ' errorfield'
        return super().form_invalid(form)


class SelectCustomerOperationListView(LoginRequiredMixin, SelectCustomerListView):
    template_name = 'minibankapp/selectcustomer_operation.html'
    paginate_by = 10

    def get_queryset(self):
        if self.request.GET.get('criteria'):
            return self.get_queryset_customer('criteria')
        else:
            return CustomerModel.objects.filter().order_by('-pk')


class SelectAcountOperationListView(LoginRequiredMixin, ListView):
    template_name = 'minibankapp/viewaccount_operation.html'

    def get_queryset(self):
        results = AccountModel.objects.filter(FK_Id_customer=self.kwargs['customer']).order_by('-pk')
        if results.exists():
            return results
        else:
            return AccountModel.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pk_customer'] = self.kwargs['customer']
        return context


""" History """
class SelectCustomerHistoryListView(LoginRequiredMixin, SelectCustomerListView):
    template_name = 'minibankapp/selectcustomer_history.html'
    paginate_by = 10

    def get_queryset(self):
        if self.request.GET.get('criteria'):
            return self.get_queryset_customer('criteria')
        else:
            return CustomerModel.objects.filter().order_by('-pk')


class SelectAcountHistoryListView(LoginRequiredMixin, ListView):
    template_name = 'minibankapp/viewaccount_history.html'

    def get_queryset(self):
        results = AccountModel.objects.filter(FK_Id_customer=self.kwargs['customer']).order_by('-pk')
        if results.exists():
            return results
        else:
            return AccountModel.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pk_customer'] = self.kwargs['customer']
        return context


class HistoryOperationListView(LoginRequiredMixin, ListView):
    template_name = 'minibankapp/historyoperation.html'
    paginate_by = 10

    def get_queryset(self):
        results = OperationModel.objects.filter(FK_Id_account=self.kwargs['account']).order_by('-Operation_date')
        if results.exists():
            return results
        else:
            return OperationModel.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pk_customer'] = self.kwargs['customer']
        context['id_account'] = self.kwargs['account']
        context['nr_iban'] = AccountModel.objects.get(pk=self.kwargs['account']).Number_IBAN
        return context


class HistoryExportListView(LoginRequiredMixin, ListView):

    def get(self, request, *args, **kwargs):
        side = Side(style='dashed', color='FF000000')
        border_around = Border(left=side, right=side, top=side, bottom=side)
        file_name = 'History_operations.xlsx'
        fields = [
                    'Id_operation',
                    'Type_operation',
                    'Value_operation',
                    'Balance_after_operation',
                    'Operation_date']
        data = OperationModel.objects.filter(FK_Id_account=self.kwargs['account']).order_by('-Operation_date').values_list(*fields)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        workbook = Workbook()
        workbook.iso_dates = True
        worksheet = workbook.active
        worksheet.title = 'Operations'
        # Column headers
        headers = [
                    'Id operation',
                    'Typ of operation',
                    'Value operation',
                    'Balance after operation',
                    'Operation date']
        for column_number, column_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=column_number)
            cell.value = column_title
            cell.font = Font(bold=True, italic=True)
            cell.fill = PatternFill(fgColor='0000FFFF', fill_type='solid')
            cell.border = border_around
        # Cell data
        for row_number, row in enumerate(data, 1):
            for column_number, cell_value in enumerate(row, 1):
                if type(cell_value) is datetime.datetime:
                    cell_value = cell_value.replace(tzinfo=None)
                    cell_value = cell_value.strftime('%d.%m.%Y %H:%M:%S')
                if column_number == 2:
                    match cell_value:
                        case 1:
                            cell_value = 'Deposit'
                        case 2:
                            cell_value = 'Withdrawal'
                        case 3:
                            cell_value = 'Interest'
                cell = worksheet.cell(row=row_number+1, column=column_number)
                cell.value = cell_value
                cell.border = border_around
                if type(cell_value) is decimal.Decimal:
                    cell.number_format = '#,##0.00'
        # AutoFit column width
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.1
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        workbook.save(response)
        return response
